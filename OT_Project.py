import openpyxl
from pypdf import PdfReader, PdfWriter
from pathlib import Path
from tkinter import Tk, filedialog
import datetime

#Current Pay Period
today = datetime.date.today()
friday = today + datetime.timedelta( (4-today.weekday()) % 7 )
next_friday = friday + datetime.timedelta(weeks=1)
current_week = datetime.date\
    (int(friday.strftime('%Y')), int(friday.strftime('%m')), int(friday.strftime('%d'))).isocalendar().week
MDY_friday = str(friday.strftime("%#m-%#d-%Y"))
MDY_next_friday = str(next_friday.strftime("%#m-%#d-%Y"))
if (current_week % 2) == 0:
    pay_period = MDY_friday
else:
    pay_period = MDY_next_friday
#Choose excel file
#Create a Tkinter root window for excel
root = Tk()
root.withdraw()
#Display a file dialog box to select the excel file
excel_path = filedialog.askopenfilename(title="Select excel file",
                                        filetypes=(("Excel Files", "*.xlsx"), ("All Files", "*.*")))
if not excel_path:
    exit()
#Display a file dialog box to select the pdf file
pdf_path = filedialog.askopenfilename(title="Select pdf file",
                                         filetypes=(("PDF Files", "*.pdf"), ("All Files", "*.*")))
if not pdf_path:
    exit()

#Opening the excel file and getting active sheet
wb_obj = openpyxl.load_workbook(excel_path)
sheet_obj = wb_obj.active
# Getting the value of maximum rows
row = sheet_obj.max_row

#Downloads file path
downloads_path = str(Path.home() / "Downloads")

#PDF reader and writer
reader = PdfReader(pdf_path)
writer = PdfWriter()
#Loop through the input and add each page
for page in range(len(reader.pages)):
    writer.add_page(reader.pages[page])
#Get the fields
fields = reader.get_fields()

#Excel Loop
for i in range(9, row + 1):
#Importing data required from excel
    cell_obj = sheet_obj.cell(row=i, column=7)
    OAKS_ID = cell_obj.value
    if OAKS_ID is not None and "Cost Center" not in str(OAKS_ID) and "SPRC Total" not in str(OAKS_ID):
        cell_obj = sheet_obj.cell(row=i, column=8)
        Employee = cell_obj.value
        cell_obj = sheet_obj.cell(row=i, column=9)
        Class = cell_obj.value
        cell_obj = sheet_obj.cell(row=i, column=18)
        Total_Cost = cell_obj.value
#Approval Amounts
        if Total_Cost >= 15000:
            Admin_Approval = "District Deputy Director over 15000 Justification as of pay period ending"
            page_number = 1
        elif Total_Cost >= 10000:
            Admin_Approval = "Administrator Approval over 10000 Justification as of pay period ending"
            page_number = 1
        else:
            Admin_Approval = "Supervisor Approval over 5000 Justification as of pay period ending"
            page_number = 0
#Filling in blanks and saving as a pdf for each employee
        writer.update_page_form_field_values(
            writer.pages[0], {"Employee _Name": Employee})
        writer.update_page_form_field_values(
            writer.pages[0], {"Classification": Class})
        writer.update_page_form_field_values(
            writer.pages[page_number], {Admin_Approval: pay_period})
        with open(downloads_path+"/"+Employee+" Overtime Form (filled-in)"+".pdf", "wb") as output_stream:
            writer.write(output_stream)